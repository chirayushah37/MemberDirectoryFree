import os
import sqlite3
import json
import http.client
import re
import calendar
import secrets
from functools import wraps
from io import BytesIO
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from urllib.parse import quote

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
try:
    import qrcode
except ImportError:
    qrcode = None
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from itsdangerous import BadSignature, URLSafeSerializer

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "bookings.db"
MEMBERS_XLSX_PATH = BASE_DIR / "ajs.xlsx"

MURTI_OPTIONS = [
    "Vasupujya Swami",
    "Neminath bhagwan",
    "Vimalnath Bhagwan",
    "Chandraprabhu swami",
]

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "replace-this-secret-key")
ADMIN_PASSWORD = "Chirayu@123123"
ADMIN_DELETE_PASSWORD = "delete"
ADMIN_EDIT_PASSWORD = "edit"
LAST_BOOKING_DATE = date(2026, 4, 30)
ADMIN_FILTER_COLUMNS = [
    "id",
    "booking_number",
    "booking_date",
    "weekday",
    "name",
    "membership_number",
    "murti",
    "address",
    "mobile_number",
    "created_at",
]
WEEKDAY_OPTIONS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
ADMIN_SORT_COLUMNS = {
    "id",
    "booking_number",
    "booking_date",
    "name",
    "membership_number",
    "murti",
    "address",
    "mobile_number",
    "created_at",
}
SANGH_ACTIVITY_AREAS = [
    "Deraser, Jinalay",
    "Ayembil shala",
    "Sadharmik Bhakti",
    "Aradhana bhuvan",
    "Pathshala",
    "Vaiya vach",
    "Vihaar",
    "Sadharmik sahay",
    "Anushthan",
]
MEMBER_DIRECTORY: dict[str, object] = {
    "entries": [],
    "by_membership": {},
}
IST = timezone(timedelta(hours=5, minutes=30), name="IST")
SQLITE_TIMESTAMP_FORMAT = "%Y-%m-%d %H:%M:%S"
OTP_SESSION_KEY = "booking_otp_state"
MEMBER_DIRECTORY_OTP_SESSION_KEY = "member_directory_otp_state"
MEMBER_DIRECTORY_AUTH_SESSION_KEY = "member_directory_mobile"
MEMBER_SEARCH_OTP_SESSION_KEY = "member_search_otp_state"
MEMBER_SEARCH_AUTH_SESSION_KEY = "member_search_mobile"
QR_CODE_SALT = "member-qrcode-v1"
OTP_EXPIRY_MINUTES = 10
SMS_API_LINK_TEMPLATE = (
    "https://chirayusoft.msg4all.com/GatewayAPI/rest?msg={message}&v=1.1&userid=2000263964"
    "&password=CEa&send_to={mobile}&msg_type=text&method=sendMessage"
)


def build_booking_number(booking_id: int) -> str:
    if booking_id > 9999:
        raise ValueError("Booking sequence exceeded 4 digits.")
    return f"{booking_id:04d}"


def now_ist() -> datetime:
    return datetime.now(IST)


def today_ist() -> date:
    return now_ist().date()


def is_tomorrow_booking_cutoff_passed(booking_day: date) -> bool:
    now = now_ist()
    tomorrow = now.date() + timedelta(days=1)
    cutoff_time = time(hour=12, minute=0)
    return booking_day == tomorrow and now.time() >= cutoff_time


def parse_created_at_timestamp(value: object) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None

    for parser in (
        lambda x: datetime.strptime(x, SQLITE_TIMESTAMP_FORMAT).replace(tzinfo=timezone.utc),
        lambda x: datetime.fromisoformat(x),
    ):
        try:
            parsed = parser(raw)
            break
        except ValueError:
            parsed = None
    else:
        return None

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed


def format_created_at_ist(value: object) -> str:
    parsed = parse_created_at_timestamp(value)
    if not parsed:
        return str(value or "")
    return parsed.astimezone(IST).strftime("%d-%b-%Y %I:%M %p IST")


def booking_row_to_dict(row: sqlite3.Row) -> dict[str, str]:
    return {
        "id": row["id"],
        "booking_number": row["booking_number"] or "",
        "booking_date": row["booking_date"],
        "name": row["name"],
        "membership_number": row["membership_number"],
        "murti": row["murti"],
        "address": row["address"],
        "mobile_number": row["mobile_number"],
        "created_at": format_created_at_ist(row["created_at"]),
    }


def generate_otp() -> str:
    return f"{secrets.randbelow(900000) + 100000}"


def clear_otp_session() -> None:
    session.pop(OTP_SESSION_KEY, None)


def get_otp_session() -> dict[str, str] | None:
    otp_state = session.get(OTP_SESSION_KEY)
    if not isinstance(otp_state, dict):
        return None
    return otp_state


def is_otp_verified_for_mobile(mobile_number: str) -> bool:
    otp_state = get_otp_session()
    if not otp_state:
        return False
    if otp_state.get("mobile_number") != mobile_number:
        return False
    if not otp_state.get("verified"):
        return False
    expires_at = otp_state.get("expires_at", "")
    try:
        expiry_time = datetime.fromisoformat(expires_at)
    except ValueError:
        return False
    if expiry_time.tzinfo is None:
        expiry_time = expiry_time.replace(tzinfo=IST)
    return expiry_time >= now_ist()


def clear_member_directory_otp_session() -> None:
    session.pop(MEMBER_DIRECTORY_OTP_SESSION_KEY, None)


def clear_member_search_otp_session() -> None:
    session.pop(MEMBER_SEARCH_OTP_SESSION_KEY, None)


def get_member_directory_otp_session() -> dict[str, str] | None:
    otp_state = session.get(MEMBER_DIRECTORY_OTP_SESSION_KEY)
    if not isinstance(otp_state, dict):
        return None
    return otp_state


def get_member_search_otp_session() -> dict[str, str] | None:
    otp_state = session.get(MEMBER_SEARCH_OTP_SESSION_KEY)
    if not isinstance(otp_state, dict):
        return None
    return otp_state


def is_member_directory_authenticated() -> bool:
    mobile_number = session.get(MEMBER_DIRECTORY_AUTH_SESSION_KEY, "")
    return isinstance(mobile_number, str) and bool(re.fullmatch(r"\d{10}", mobile_number))


def is_member_search_authenticated() -> bool:
    mobile_number = session.get(MEMBER_SEARCH_AUTH_SESSION_KEY, "")
    return isinstance(mobile_number, str) and bool(re.fullmatch(r"\d{10}", mobile_number))


def get_authenticated_member_mobile() -> str:
    mobile_number = session.get(MEMBER_DIRECTORY_AUTH_SESSION_KEY, "")
    if isinstance(mobile_number, str):
        return mobile_number
    return ""


def get_authenticated_member_search_mobile() -> str:
    mobile_number = session.get(MEMBER_SEARCH_AUTH_SESSION_KEY, "")
    if isinstance(mobile_number, str):
        return mobile_number
    return ""


def send_otp_sms(mobile_number: str, otp: str) -> None:
    message = f"Your OTP is {otp}\nCHIRAYU SOFTWARE"
    url = SMS_API_LINK_TEMPLATE.format(message=quote(message), mobile=quote(mobile_number))
    conn = http.client.HTTPSConnection("chirayusoft.msg4all.com", timeout=30)
    try:
        conn.request("GET", url.replace("https://chirayusoft.msg4all.com", ""))
        response = conn.getresponse()
        response_body = response.read().decode("utf-8", errors="replace")
        if response.status >= 400:
            raise RuntimeError(f"SMS API error {response.status}: {response_body}")
    finally:
        conn.close()


def normalize_text(value: object) -> str:
    return " ".join(str(value or "").upper().split())


def format_excel_part(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_mobile(value: object) -> str:
    matches = re.findall(r"\d{10}", str(value or ""))
    return matches[0] if matches else ""


def load_member_directory() -> None:
    entries: list[dict[str, str]] = []
    by_membership: dict[str, dict[str, str]] = {}
    if not MEMBERS_XLSX_PATH.exists():
        MEMBER_DIRECTORY["entries"] = entries
        MEMBER_DIRECTORY["by_membership"] = by_membership
        return

    workbook = load_workbook(MEMBERS_XLSX_PATH, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            part_a = format_excel_part(row[0] if len(row) > 0 else "")
            part_b = format_excel_part(row[1] if len(row) > 1 else "")
            membership_number = f"{part_a}{part_b}".strip()
            if not membership_number:
                continue

            name = format_excel_part(row[2] if len(row) > 2 else "")
            father_name = format_excel_part(row[3] if len(row) > 3 else "")
            surname = format_excel_part(row[4] if len(row) > 4 else "")
            full_name = " ".join([x for x in [name, father_name, surname] if x]).strip()
            if not full_name:
                continue

            address = format_excel_part(row[5] if len(row) > 5 else "")
            mobile_number = clean_mobile(row[7] if len(row) > 7 else "")

            entry = {
                "membership_number": membership_number,
                "name": full_name,
                "address": address,
                "mobile_number": mobile_number,
                "search_blob": normalize_text(f"{membership_number} {full_name} {address} {mobile_number}"),
            }
            entries.append(entry)
            by_membership[membership_number.upper()] = entry
    finally:
        workbook.close()

    entries.sort(key=lambda x: x["name"])
    MEMBER_DIRECTORY["entries"] = entries
    MEMBER_DIRECTORY["by_membership"] = by_membership


def search_members(query: str, limit: int = 25) -> list[dict[str, str]]:
    entries = MEMBER_DIRECTORY.get("entries", [])
    if not isinstance(entries, list):
        return []
    normalized_query = normalize_text(query)
    if not normalized_query:
        selected = entries[:limit]
    else:
        matched = [entry for entry in entries if normalized_query in entry["search_blob"]]
        matched.sort(
            key=lambda x: (
                not x["name"].upper().startswith(normalized_query),
                not x["membership_number"].upper().startswith(normalized_query),
                x["name"],
            )
        )
        selected = matched[:limit]

    return [
        {
            "membership_number": entry["membership_number"],
            "name": entry["name"],
            "address": entry["address"],
            "mobile_number": entry["mobile_number"],
        }
        for entry in selected
    ]


def find_member_by_name(name: str) -> dict[str, str] | None:
    normalized_name = normalize_text(name)
    if not normalized_name:
        return None
    entries = MEMBER_DIRECTORY.get("entries", [])
    if not isinstance(entries, list):
        return None
    for entry in entries:
        if normalize_text(entry["name"]) == normalized_name:
            return entry
    return None


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=15, isolation_level=None)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA busy_timeout=15000;")
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn


def init_db() -> None:
    schema = """
    CREATE TABLE IF NOT EXISTS bookings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        booking_date TEXT NOT NULL,
        name TEXT NOT NULL CHECK (trim(name) <> ''),
        membership_number TEXT NOT NULL CHECK (trim(membership_number) <> ''),
        murti TEXT NOT NULL CHECK (trim(murti) <> ''),
        address TEXT NOT NULL CHECK (trim(address) <> ''),
        mobile_number TEXT NOT NULL CHECK (trim(mobile_number) <> ''),
        is_dummy INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(booking_date, murti)
    );

    DROP TRIGGER IF EXISTS bookings_limit_per_day;
    CREATE TRIGGER bookings_limit_per_day
    BEFORE INSERT ON bookings
    FOR EACH ROW
    WHEN (SELECT COUNT(1) FROM bookings WHERE booking_date = NEW.booking_date) >= 2
    BEGIN
        SELECT RAISE(ABORT, 'A maximum of two bookings are allowed for this date.');
    END;

    DROP TRIGGER IF EXISTS bookings_block_sunday;
    CREATE TRIGGER bookings_block_sunday
    BEFORE INSERT ON bookings
    FOR EACH ROW
    WHEN strftime('%w', NEW.booking_date) = '0' AND COALESCE(NEW.is_dummy, 0) = 0
    BEGIN
        SELECT RAISE(ABORT, 'Booking is already done');
    END;

    CREATE TABLE IF NOT EXISTS member_registrations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mobile_number TEXT NOT NULL UNIQUE CHECK (trim(mobile_number) <> ''),
        approval_status TEXT NOT NULL DEFAULT 'pending' CHECK (approval_status IN ('pending', 'approved')),
        head_of_family_details TEXT NOT NULL DEFAULT '{}',
        family_members_details TEXT NOT NULL DEFAULT '[]',
        sangh_activity_interest TEXT NOT NULL DEFAULT '{}',
        business_details TEXT NOT NULL DEFAULT '{}',
        fees_details TEXT NOT NULL DEFAULT '[]',
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        approved_at TEXT
    );

    CREATE TABLE IF NOT EXISTS site_settings (
        setting_key TEXT PRIMARY KEY,
        setting_value TEXT NOT NULL DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS new_membership_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL CHECK (trim(name) <> ''),
        mobile_number TEXT NOT NULL CHECK (trim(mobile_number) <> ''),
        reference_name TEXT NOT NULL CHECK (trim(reference_name) <> ''),
        reference_membership_number TEXT NOT NULL CHECK (trim(reference_membership_number) <> ''),
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    """
    conn = get_db_connection()
    try:
        conn.executescript(schema)
        table_info = conn.execute("PRAGMA table_info(bookings)").fetchall()
        column_names = {row["name"] for row in table_info}
        if "booking_number" not in column_names:
            conn.execute("ALTER TABLE bookings ADD COLUMN booking_number TEXT")
        if "is_dummy" not in column_names:
            conn.execute("ALTER TABLE bookings ADD COLUMN is_dummy INTEGER NOT NULL DEFAULT 0")
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_bookings_booking_number ON bookings(booking_number)")
        member_table_info = conn.execute("PRAGMA table_info(member_registrations)").fetchall()
        member_column_names = {row["name"] for row in member_table_info}
        if member_table_info:
            if "approved_at" not in member_column_names:
                conn.execute("ALTER TABLE member_registrations ADD COLUMN approved_at TEXT")
            if "sangh_activity_interest" not in member_column_names:
                conn.execute("ALTER TABLE member_registrations ADD COLUMN sangh_activity_interest TEXT NOT NULL DEFAULT '{}'")
            if "fees_details" not in member_column_names:
                conn.execute("ALTER TABLE member_registrations ADD COLUMN fees_details TEXT NOT NULL DEFAULT '[]'")
        rows_to_update = conn.execute("SELECT id FROM bookings").fetchall()
        for row in rows_to_update:
            conn.execute(
                "UPDATE bookings SET booking_number = ? WHERE id = ?",
                (build_booking_number(int(row["id"])), int(row["id"])),
            )
    finally:
        conn.close()


def get_bookings() -> list[sqlite3.Row]:
    conn = get_db_connection()
    try:
        return conn.execute(
            """
            SELECT booking_date, murti
            FROM bookings
            ORDER BY booking_date DESC, murti ASC
            """
        ).fetchall()
    finally:
        conn.close()


def get_booking_counts_by_date() -> dict[str, int]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT booking_date, COUNT(1) AS booking_count
            FROM bookings
            GROUP BY booking_date
            """
        ).fetchall()
    finally:
        conn.close()
    return {row["booking_date"]: int(row["booking_count"]) for row in rows}


def get_site_setting(setting_key: str, default: str = "") -> str:
    conn = get_db_connection()
    try:
        row = conn.execute(
            "SELECT setting_value FROM site_settings WHERE setting_key = ?",
            (setting_key,),
        ).fetchone()
    finally:
        conn.close()
    return str(row["setting_value"]) if row else default


def update_site_setting(setting_key: str, setting_value: str) -> None:
    conn = get_db_connection()
    try:
        conn.execute(
            """
            INSERT INTO site_settings (setting_key, setting_value)
            VALUES (?, ?)
            ON CONFLICT(setting_key) DO UPDATE SET setting_value = excluded.setting_value
            """,
            (setting_key, setting_value),
        )
    finally:
        conn.close()


@app.context_processor
def inject_site_settings():
    return {"marquee_text": get_site_setting("marquee_text")}


def insert_new_membership_request(form_data: dict[str, str]) -> None:
    conn = get_db_connection()
    try:
        conn.execute(
            """
            INSERT INTO new_membership_requests
                (name, mobile_number, reference_name, reference_membership_number)
            VALUES (?, ?, ?, ?)
            """,
            (
                form_data["name"],
                form_data["mobile_number"],
                form_data["reference_name"],
                form_data["reference_membership_number"],
            ),
        )
    finally:
        conn.close()


def get_new_membership_requests() -> list[dict[str, str]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, name, mobile_number, reference_name, reference_membership_number, created_at
            FROM new_membership_requests
            ORDER BY created_at DESC, id DESC
            """
        ).fetchall()
    finally:
        conn.close()

    return [
        {
            "id": row["id"],
            "name": row["name"],
            "mobile_number": row["mobile_number"],
            "reference_name": row["reference_name"],
            "reference_membership_number": row["reference_membership_number"],
            "created_at": format_created_at_ist(row["created_at"]),
        }
        for row in rows
    ]


def get_advertise_member_profiles() -> list[dict[str, str]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, updated_at, approved_at
            FROM member_registrations
            ORDER BY updated_at DESC, id DESC
            """
        ).fetchall()
    finally:
        conn.close()

    profiles: list[dict[str, str]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        if str(head.get("would_like_to_advertise", "0")).strip() != "1":
            continue
        profiles.append(
            {
                "id": row["id"],
                "approval_status": row["approval_status"],
                "full_name": str(head.get("full_name", "")).strip(),
                "membership_number": str(head.get("membership_number", "")).strip(),
                "mobile_number": str(head.get("mobile_number", "")).strip() or row["mobile_number"],
                "phone": str(head.get("phone", "")).strip(),
                "occupation": str(head.get("occupation", "")).strip(),
                "address": str(head.get("address", "")).strip(),
                "city": str(head.get("city", "")).strip(),
                "email": str(head.get("email", "")).strip(),
                "updated_at": format_created_at_ist(row["updated_at"]),
                "approved_at": format_created_at_ist(row["approved_at"]) if row["approved_at"] else "",
            }
        )
    return profiles


def get_discount_interest_member_profiles() -> list[dict[str, str]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, updated_at, approved_at
            FROM member_registrations
            ORDER BY updated_at DESC, id DESC
            """
        ).fetchall()
    finally:
        conn.close()

    profiles: list[dict[str, str]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        if str(head.get("sangh_member_discount_interest", "0")).strip() != "1":
            continue
        profiles.append(
            {
                "id": row["id"],
                "approval_status": row["approval_status"],
                "full_name": str(head.get("full_name", "")).strip(),
                "membership_number": str(head.get("membership_number", "")).strip(),
                "mobile_number": str(head.get("mobile_number", "")).strip() or row["mobile_number"],
                "phone": str(head.get("phone", "")).strip(),
                "occupation": str(head.get("occupation", "")).strip(),
                "address": str(head.get("address", "")).strip(),
                "city": str(head.get("city", "")).strip(),
                "email": str(head.get("email", "")).strip(),
                "updated_at": format_created_at_ist(row["updated_at"]),
                "approved_at": format_created_at_ist(row["approved_at"]) if row["approved_at"] else "",
            }
        )
    return profiles


def get_pathshala_interest_member_profiles() -> list[dict[str, str]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, family_members_details, updated_at, approved_at
            FROM member_registrations
            ORDER BY updated_at DESC, id DESC
            """
        ).fetchall()
    finally:
        conn.close()

    profiles: list[dict[str, str]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        head_name = str(head.get("full_name", "")).strip()
        membership_number = str(head.get("membership_number", "")).strip()
        contact_mobile = str(head.get("mobile_number", "")).strip() or row["mobile_number"]
        for member in parse_json_array(row["family_members_details"]):
            if str(member.get("pathshala_interest", "0")).strip() != "1":
                continue
            birthdate = str(member.get("birthdate", "")).strip()
            calculated_age = calculate_age_from_birthdate(birthdate)
            profiles.append(
                {
                    "registration_id": row["id"],
                    "approval_status": row["approval_status"],
                    "head_of_family": head_name,
                    "membership_number": membership_number,
                    "contact_mobile": contact_mobile,
                    "name": str(member.get("name", "")).strip(),
                    "relation": str(member.get("relation", "")).strip(),
                    "sex": str(member.get("sex", "")).strip(),
                    "birthdate": birthdate,
                    "age": "" if calculated_age is None else str(calculated_age),
                    "mobile_number": str(member.get("mobile_number", "")).strip() or contact_mobile,
                    "education": str(member.get("education", "")).strip(),
                    "updated_at": format_created_at_ist(row["updated_at"]),
                    "approved_at": format_created_at_ist(row["approved_at"]) if row["approved_at"] else "",
                }
            )
    return profiles


def get_sangh_activity_interest_profiles(selected_area: str = "") -> list[dict[str, str]]:
    selected_area = selected_area.strip()
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, family_members_details,
                   sangh_activity_interest, updated_at, approved_at
            FROM member_registrations
            ORDER BY updated_at DESC, id DESC
            """
        ).fetchall()
    finally:
        conn.close()

    def selected_areas_from(activity: object) -> list[str]:
        if not isinstance(activity, dict):
            return []
        raw_areas = activity.get("area_interested", [])
        if not isinstance(raw_areas, list):
            return []
        return [str(area).strip() for area in raw_areas if str(area).strip()]

    profiles: list[dict[str, str]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        family_members = parse_json_array(row["family_members_details"])
        activity = normalize_sangh_activity_interest(parse_json_object(row["sangh_activity_interest"]), family_members)

        head_name = str(head.get("full_name", "")).strip()
        membership_number = str(head.get("membership_number", "")).strip()
        contact_mobile = str(head.get("mobile_number", "")).strip() or row["mobile_number"]
        base = {
            "registration_id": row["id"],
            "approval_status": row["approval_status"],
            "head_of_family": head_name,
            "membership_number": membership_number,
            "contact_mobile": contact_mobile,
            "updated_at": format_created_at_ist(row["updated_at"]),
            "approved_at": format_created_at_ist(row["approved_at"]) if row["approved_at"] else "",
        }

        head_activity = activity.get("head_of_family", {})
        head_areas = selected_areas_from(head_activity)
        if head_areas and (not selected_area or selected_area in head_areas):
            profiles.append(
                {
                    **base,
                    "name": head_name,
                    "person_type": "Head Of Family",
                    "relation": "Head Of Family",
                    "mobile_number": contact_mobile,
                    "occupation": str(head.get("occupation", "")).strip(),
                    "how_can_help": str(head_activity.get("how_can_help", "")).strip() if isinstance(head_activity, dict) else "",
                    "weekly_hours": str(head_activity.get("weekly_hours", "")).strip() if isinstance(head_activity, dict) else "",
                    "area_interested": ", ".join(head_areas),
                }
            )

        family_activity = activity.get("family_members", [])
        for index, member in enumerate(family_members):
            member_activity = family_activity[index] if isinstance(family_activity, list) and index < len(family_activity) else {}
            member_areas = selected_areas_from(member_activity)
            if not member_areas or (selected_area and selected_area not in member_areas):
                continue
            profiles.append(
                {
                    **base,
                    "name": str(member.get("name", "")).strip(),
                    "person_type": "Family Member",
                    "relation": str(member.get("relation", "")).strip(),
                    "mobile_number": str(member.get("mobile_number", "")).strip() or contact_mobile,
                    "occupation": str(member.get("occupation", "")).strip(),
                    "how_can_help": str(member_activity.get("how_can_help", "")).strip() if isinstance(member_activity, dict) else "",
                    "weekly_hours": str(member_activity.get("weekly_hours", "")).strip() if isinstance(member_activity, dict) else "",
                    "area_interested": ", ".join(member_areas),
                }
            )
    return profiles


def booking_exists_for_date_and_murti(booking_date: str, murti: str) -> bool:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT 1
            FROM bookings
            WHERE booking_date = ? AND murti = ?
            LIMIT 1
            """,
            (booking_date, murti),
        ).fetchone()
        return row is not None
    except sqlite3.OperationalError:
        return False
    finally:
        conn.close()


def booking_exists_for_date_and_murti_excluding_id(booking_date: str, murti: str, booking_id: int) -> bool:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT 1
            FROM bookings
            WHERE booking_date = ? AND murti = ? AND id <> ?
            LIMIT 1
            """,
            (booking_date, murti, booking_id),
        ).fetchone()
        return row is not None
    except sqlite3.OperationalError:
        return False
    finally:
        conn.close()


def count_bookings_for_date_excluding_id(booking_date: str, booking_id: int) -> int:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT COUNT(1) AS booking_count
            FROM bookings
            WHERE booking_date = ? AND id <> ?
            """,
            (booking_date, booking_id),
        ).fetchone()
        return int(row["booking_count"] if row else 0)
    finally:
        conn.close()


def booking_number_exists_excluding_id(booking_number: str, booking_id: int) -> bool:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT 1
            FROM bookings
            WHERE booking_number = ? AND id <> ?
            LIMIT 1
            """,
            (booking_number, booking_id),
        ).fetchone()
        return row is not None
    finally:
        conn.close()


def get_calendar_months_view(year: int, months: list[int]) -> list[dict[str, object]]:
    counts_by_date = get_booking_counts_by_date()
    cal = calendar.Calendar(firstweekday=0)  # Monday
    month_views: list[dict[str, object]] = []

    for month in months:
        weeks: list[list[dict[str, object]]] = []
        for week in cal.monthdayscalendar(year, month):
            week_days: list[dict[str, object]] = []
            for day in week:
                if day == 0:
                    week_days.append({"in_month": False, "day": "", "count": 0, "date_iso": ""})
                    continue
                date_iso = f"{year}-{month:02d}-{day:02d}"
                count = counts_by_date.get(date_iso, 0)
                week_days.append({"in_month": True, "day": day, "count": count, "date_iso": date_iso})
            weeks.append(week_days)
        month_views.append(
            {
                "month_name": datetime(year, month, 1).strftime("%B %Y"),
                "weeks": weeks,
            }
        )
    return month_views


def get_bookings_for_display() -> list[dict[str, str]]:
    display_rows: list[dict[str, str]] = []
    for booking in get_bookings():
        booking_date = booking["booking_date"]
        try:
            dt = datetime.strptime(booking_date, "%Y-%m-%d")
            display_date = dt.strftime("%d-%b (%A)")
        except ValueError:
            display_date = booking_date
        display_rows.append({"display_date": display_date, "murti": booking["murti"]})
    return display_rows


def render_index_page(form_data: dict[str, str]):
    return render_template(
        "index.html",
        murtis=MURTI_OPTIONS,
        bookings=get_bookings_for_display(),
        form_data=form_data,
        today_date=today_ist().isoformat(),
        last_booking_date=LAST_BOOKING_DATE.isoformat(),
        calendar_months=get_calendar_months_view(LAST_BOOKING_DATE.year, [3, 4]),
    )


def get_booking(booking_id: int) -> dict[str, str] | None:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT id, booking_number, booking_date, name, membership_number, murti, address, mobile_number, created_at
            FROM bookings
            WHERE id = ?
            """,
            (booking_id,),
        ).fetchone()
        if not row:
            return None
        return booking_row_to_dict(row)
    finally:
        conn.close()


def get_booking_for_edit(booking_id: int) -> dict[str, str] | None:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT id, booking_number, booking_date, name, membership_number, murti, address, mobile_number, created_at, is_dummy
            FROM bookings
            WHERE id = ?
            """,
            (booking_id,),
        ).fetchone()
        if not row:
            return None
        return {
            "id": row["id"],
            "booking_number": row["booking_number"] or "",
            "booking_date": row["booking_date"],
            "name": row["name"],
            "membership_number": row["membership_number"],
            "murti": row["murti"],
            "address": row["address"],
            "mobile_number": row["mobile_number"],
            "created_at": format_created_at_ist(row["created_at"]),
            "is_dummy": str(row["is_dummy"] or 0),
        }
    finally:
        conn.close()


def normalize_filters(args: dict[str, str]) -> dict[str, str]:
    filters: dict[str, str] = {}
    for column in ADMIN_FILTER_COLUMNS:
        filters[column] = args.get(column, "").strip()
    return filters


def normalize_admin_sort(args: dict[str, str]) -> tuple[str, str]:
    sort_by = args.get("sort_by", "booking_date").strip()
    sort_dir = args.get("sort_dir", "desc").strip().lower()
    if sort_by not in ADMIN_SORT_COLUMNS:
        sort_by = "booking_date"
    if sort_dir not in {"asc", "desc"}:
        sort_dir = "desc"
    return sort_by, sort_dir


def format_admin_date(booking_date: str) -> tuple[str, str, date | None]:
    try:
        dt = datetime.strptime(booking_date, "%Y-%m-%d")
        return dt.strftime("%d-%b"), dt.strftime("%A"), dt.date()
    except ValueError:
        return booking_date, "", None


def get_admin_bookings(filters: dict[str, str], sort_by: str, sort_dir: str) -> list[dict[str, str]]:
    where_clauses: list[str] = []
    params: list[str] = []

    for column, value in filters.items():
        if not value:
            continue
        if column in {"booking_date", "weekday", "created_at"}:
            continue
        if column == "id":
            where_clauses.append("CAST(id AS TEXT) LIKE ?")
        else:
            where_clauses.append(f"{column} LIKE ?")
        params.append(f"%{value}%")

    query = """
        SELECT id, booking_number, booking_date, name, membership_number, murti, address, mobile_number, created_at
        FROM bookings
    """
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
    query += " ORDER BY booking_date DESC, created_at DESC"

    conn = get_db_connection()
    try:
        rows = conn.execute(query, params).fetchall()
    finally:
        conn.close()

    bookings: list[dict[str, str]] = []
    date_filter = filters.get("booking_date", "").upper()
    weekday_filter = filters.get("weekday", "")

    for row in rows:
        date_display, weekday, parsed_date = format_admin_date(row["booking_date"])
        created_at_display = format_created_at_ist(row["created_at"])
        parsed_created_at = parse_created_at_timestamp(row["created_at"])
        if date_filter and date_filter not in date_display.upper():
            continue
        if weekday_filter and weekday_filter != weekday:
            continue
        created_at_filter = filters.get("created_at", "").upper()
        if created_at_filter and created_at_filter not in created_at_display.upper():
            continue

        bookings.append(
            {
                "id": row["id"],
                "booking_number": row["booking_number"] or "",
                "booking_date": row["booking_date"],
                "booking_date_display": date_display,
                "weekday": weekday,
                "name": row["name"],
                "membership_number": row["membership_number"],
                "murti": row["murti"],
                "address": row["address"],
                "mobile_number": row["mobile_number"],
                "created_at": created_at_display,
                "_sort_date": parsed_date,
                "_sort_created_at": parsed_created_at or datetime.min.replace(tzinfo=timezone.utc),
            }
        )

    reverse = sort_dir == "desc"

    def sort_key(booking: dict[str, str]):
        if sort_by == "id":
            return int(booking["id"])
        if sort_by == "booking_date":
            return booking["_sort_date"] or date.min
        if sort_by == "created_at":
            return booking["_sort_created_at"]
        return str(booking.get(sort_by, "")).upper()

    bookings.sort(key=sort_key, reverse=reverse)
    for booking in bookings:
        booking.pop("_sort_date", None)
        booking.pop("_sort_created_at", None)
    return bookings


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("is_admin_authenticated"):
            return redirect(url_for("admin_login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapped


def member_directory_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not is_member_directory_authenticated():
            flash("Please login with mobile number and OTP first.", "error")
            return redirect(url_for("member_directory_login"))
        return view_func(*args, **kwargs)

    return wrapped


def member_search_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not is_member_search_authenticated():
            flash("Please login with mobile number and OTP first.", "error")
            return redirect(url_for("member_search_login", registration_id=kwargs.get("registration_id")))
        return view_func(*args, **kwargs)

    return wrapped


def parse_json_object(raw_value: object) -> dict[str, object]:
    try:
        parsed = json.loads(str(raw_value or "{}"))
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def parse_json_array(raw_value: object) -> list[dict[str, object]]:
    try:
        parsed = json.loads(str(raw_value or "[]"))
    except json.JSONDecodeError:
        return []
    if not isinstance(parsed, list):
        return []
    return [item for item in parsed if isinstance(item, dict)]


def get_member_registration_by_mobile(mobile_number: str) -> dict[str, object] | None:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, family_members_details,
                   sangh_activity_interest, business_details, fees_details, created_at, updated_at, approved_at
            FROM member_registrations
            WHERE mobile_number = ?
            """,
            (mobile_number,),
        ).fetchone()
    finally:
        conn.close()

    if not row:
        return None

    return {
        "id": row["id"],
        "mobile_number": row["mobile_number"],
        "approval_status": row["approval_status"],
        "head_of_family_details": parse_json_object(row["head_of_family_details"]),
        "family_members_details": parse_json_array(row["family_members_details"]),
        "sangh_activity_interest": parse_json_object(row["sangh_activity_interest"]),
        "business_details": parse_json_object(row["business_details"]),
        "fees_details": parse_json_array(row["fees_details"]),
        "created_at": format_created_at_ist(row["created_at"]),
        "updated_at": format_created_at_ist(row["updated_at"]),
        "approved_at": format_created_at_ist(row["approved_at"]) if row["approved_at"] else "",
    }


def get_member_registration_by_id(registration_id: int) -> dict[str, object] | None:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, family_members_details,
                   sangh_activity_interest, business_details, fees_details, created_at, updated_at, approved_at
            FROM member_registrations
            WHERE id = ?
            """,
            (registration_id,),
        ).fetchone()
    finally:
        conn.close()

    if not row:
        return None

    return {
        "id": row["id"],
        "mobile_number": row["mobile_number"],
        "approval_status": row["approval_status"],
        "head_of_family_details": parse_json_object(row["head_of_family_details"]),
        "family_members_details": parse_json_array(row["family_members_details"]),
        "sangh_activity_interest": parse_json_object(row["sangh_activity_interest"]),
        "business_details": parse_json_object(row["business_details"]),
        "fees_details": parse_json_array(row["fees_details"]),
        "created_at": format_created_at_ist(row["created_at"]),
        "updated_at": format_created_at_ist(row["updated_at"]),
        "approved_at": format_created_at_ist(row["approved_at"]) if row["approved_at"] else "",
    }


def ensure_member_registration(mobile_number: str) -> dict[str, object]:
    existing = get_member_registration_by_mobile(mobile_number)
    if existing:
        return existing

    conn = get_db_connection()
    try:
        conn.execute(
            """
            INSERT INTO member_registrations (mobile_number)
            VALUES (?)
            """,
            (mobile_number,),
        )
    finally:
        conn.close()
    return get_member_registration_by_mobile(mobile_number) or {
        "id": 0,
        "mobile_number": mobile_number,
        "approval_status": "pending",
        "head_of_family_details": {},
        "family_members_details": [],
        "sangh_activity_interest": {},
        "business_details": {},
        "fees_details": [],
        "created_at": "",
        "updated_at": "",
        "approved_at": "",
    }


def update_member_registration_section(mobile_number: str, section: str, payload: object) -> None:
    if section not in {"head_of_family_details", "family_members_details", "sangh_activity_interest", "business_details"}:
        raise ValueError("Invalid section.")

    serialized_payload = json.dumps(payload, ensure_ascii=True)
    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE;")
        conn.execute(
            f"""
            INSERT INTO member_registrations (mobile_number, {section}, approval_status, updated_at, approved_at)
            VALUES (?, ?, 'pending', CURRENT_TIMESTAMP, NULL)
            ON CONFLICT(mobile_number) DO UPDATE SET
                {section} = excluded.{section},
                approval_status = 'pending',
                updated_at = CURRENT_TIMESTAMP,
                approved_at = NULL
            """,
            (mobile_number, serialized_payload),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def get_all_member_registrations() -> list[dict[str, object]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, family_members_details,
                   sangh_activity_interest, business_details, fees_details, created_at, updated_at, approved_at
            FROM member_registrations
            ORDER BY
                CASE approval_status WHEN 'pending' THEN 0 ELSE 1 END,
                updated_at DESC,
                created_at DESC
            """
        ).fetchall()
    finally:
        conn.close()

    registrations: list[dict[str, object]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        family_members = parse_json_array(row["family_members_details"])
        business = parse_json_object(row["business_details"])
        fees_details = parse_json_array(row["fees_details"])
        registrations.append(
            {
                "id": row["id"],
                "mobile_number": row["mobile_number"],
                "approval_status": row["approval_status"],
                "head_name": str(head.get("full_name", "")).strip(),
                "membership_number": str(head.get("membership_number", "")).strip(),
                "city": str(head.get("city", "")).strip(),
                "show_in_matrimony": any(
                    str(member.get("show_in_matrimony", "0")).strip() == "1" for member in family_members
                ),
                "business_name": str(business.get("business_name", "")).strip(),
                "fees_details": fees_details,
                "family_count": len(family_members),
                "created_at": format_created_at_ist(row["created_at"]),
                "updated_at": format_created_at_ist(row["updated_at"]),
                "approved_at": format_created_at_ist(row["approved_at"]) if row["approved_at"] else "",
            }
        )
    return registrations


def set_member_registration_approval(registration_id: int, status: str) -> bool:
    if status not in {"pending", "approved"}:
        raise ValueError("Invalid approval status.")

    approved_at = "CURRENT_TIMESTAMP" if status == "approved" else "NULL"
    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE;")
        cursor = conn.execute(
            f"""
            UPDATE member_registrations
            SET approval_status = ?, updated_at = CURRENT_TIMESTAMP, approved_at = {approved_at}
            WHERE id = ?
            """,
            (status, registration_id),
        )
        conn.commit()
        return cursor.rowcount > 0
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def add_member_registration_fee(registration_id: int, year: str, amount: str) -> bool:
    year = year.strip()
    amount = amount.strip()
    if not year or not amount:
        raise ValueError("Year and amount are required.")

    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE;")
        row = conn.execute(
            "SELECT fees_details FROM member_registrations WHERE id = ?",
            (registration_id,),
        ).fetchone()
        if not row:
            conn.rollback()
            return False

        fees_details = parse_json_array(row["fees_details"])
        fees_details.append({"year": year, "amount": amount})
        conn.execute(
            """
            UPDATE member_registrations
            SET fees_details = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (json.dumps(fees_details, ensure_ascii=True), registration_id),
        )
        conn.commit()
        return True
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def get_qr_serializer() -> URLSafeSerializer:
    return URLSafeSerializer(app.secret_key, salt=QR_CODE_SALT)


def build_member_qr_token(registration_id: int) -> str:
    return get_qr_serializer().dumps({"registration_id": int(registration_id)})


def get_registration_id_from_qr_token(token: str) -> int | None:
    try:
        payload = get_qr_serializer().loads(token)
    except BadSignature:
        return None
    if not isinstance(payload, dict):
        return None
    try:
        return int(payload.get("registration_id", 0))
    except (TypeError, ValueError):
        return None


def get_fee_verification_status(registration: dict[str, object]) -> dict[str, object]:
    current_year = now_ist().year
    accepted_years = {str(current_year), str(current_year - 1)}
    matched_fees: list[dict[str, str]] = []
    for fee in registration.get("fees_details", []):
        if not isinstance(fee, dict):
            continue
        fee_year = str(fee.get("year", "")).strip()
        fee_amount = str(fee.get("amount", "")).strip()
        if fee_amount and any(year in fee_year for year in accepted_years):
            matched_fees.append({"year": fee_year, "amount": fee_amount})

    return {
        "current_year": current_year,
        "previous_year": current_year - 1,
        "is_approved": registration.get("approval_status") == "approved",
        "has_recent_fee": bool(matched_fees),
        "matched_fees": matched_fees,
    }


def parse_head_of_family_form(form) -> dict[str, str]:
    return {
        "full_name": form.get("full_name", "").strip(),
        "membership_number": form.get("membership_number", "").strip(),
        "sex": form.get("sex", "").strip(),
        "birthdate": form.get("birthdate", "").strip(),
        "mobile_number": form.get("mobile_number", "").strip(),
        "phone": form.get("phone", "").strip(),
        "blood_group": form.get("blood_group", "").strip(),
        "education": form.get("education", "").strip(),
        "occupation": form.get("occupation", "").strip(),
        "address": form.get("address", "").strip(),
        "city": form.get("city", "").strip(),
        "pincode": form.get("pincode", "").strip(),
        "email": form.get("email", "").strip(),
        "would_like_to_advertise": "1" if form.get("would_like_to_advertise", "").strip() == "1" else "0",
        "sangh_member_discount_interest": "1" if form.get("sangh_member_discount_interest", "").strip() == "1" else "0",
    }


def parse_family_member_form(form) -> dict[str, str]:
    sex = form.get("family_member_sex", "").strip()
    birthdate = form.get("family_member_birthdate", "").strip()
    calculated_age = calculate_age_from_birthdate(birthdate)
    return {
        "name": form.get("family_member_name", "").strip(),
        "relation": form.get("family_member_relation", "").strip(),
        "sex": sex,
        "birthdate": birthdate,
        "mobile_number": form.get("family_member_mobile_number", "").strip(),
        "blood_group": form.get("family_member_blood_group", "").strip(),
        "education": form.get("family_member_education", "").strip(),
        "occupation": form.get("family_member_occupation", "").strip(),
        "married": (
            "1"
            if calculated_age is not None and calculated_age > 18 and form.get("family_member_married", "").strip() == "1"
            else "0"
        ),
        "show_in_matrimony": "1" if form.get("family_member_show_in_matrimony", "").strip() == "1" else "0",
        "mahila_mandal_interest": (
            "1"
            if sex == "Female" and form.get("family_member_mahila_mandal_interest", "").strip() == "1"
            else "0"
        ),
        "pathshala_interest": (
            "1"
            if calculated_age is not None and calculated_age < 18 and form.get("family_member_pathshala_interest", "").strip() == "1"
            else "0"
        ),
        "sangh_activity_interest": parse_sangh_activity_interest_form(
            form,
            area_field_name="family_member_area_interested",
            how_can_help_field_name="family_member_how_can_help",
            weekly_hours_field_name="family_member_weekly_hours",
        ),
    }


def parse_business_details_form(form) -> dict[str, str]:
    return {
        "business_name": form.get("business_name", "").strip(),
        "business_type": form.get("business_type", "").strip(),
        "designation": form.get("designation", "").strip(),
        "business_contact_mobile": form.get("business_contact_mobile", "").strip(),
        "office_address": form.get("office_address", "").strip(),
        "website": form.get("website", "").strip(),
        "notes": form.get("notes", "").strip(),
        "show_in_search": "1" if form.get("show_in_search", "").strip() == "1" else "0",
    }


def parse_sangh_activity_interest_form(
    form,
    *,
    area_field_name: str = "area_interested",
    how_can_help_field_name: str = "how_can_help",
    weekly_hours_field_name: str = "weekly_hours",
) -> dict[str, object]:
    selected_areas = [item.strip() for item in form.getlist(area_field_name) if item.strip()]
    return {
        "how_can_help": form.get(how_can_help_field_name, "").strip(),
        "weekly_hours": form.get(weekly_hours_field_name, "").strip(),
        "area_interested": selected_areas,
    }


def calculate_age_from_birthdate(birthdate: str) -> int | None:
    try:
        birthdate_value = date.fromisoformat(str(birthdate).strip())
    except ValueError:
        return None

    today = now_ist().date()
    age = today.year - birthdate_value.year
    if (today.month, today.day) < (birthdate_value.month, birthdate_value.day):
        age -= 1
    return age if age >= 0 else None


def normalize_sangh_activity_interest(
    raw_interest: dict[str, object], family_members: list[dict[str, str]]
) -> dict[str, object]:
    normalized: dict[str, object] = {
        "head_of_family": {},
        "family_members": [{} for _ in family_members],
    }

    stored_family = raw_interest.get("family_members", [])
    if isinstance(stored_family, list):
        for index, entry in enumerate(stored_family[: len(family_members)]):
            if isinstance(entry, dict):
                normalized["family_members"][index] = entry

    stored_head = raw_interest.get("head_of_family")
    if isinstance(stored_head, dict):
        normalized["head_of_family"] = stored_head

    legacy_keys = {"how_can_help", "weekly_hours", "area_interested"}
    if (
        isinstance(raw_interest, dict)
        and not normalized["head_of_family"]
        and any(key in raw_interest for key in legacy_keys)
    ):
        normalized["head_of_family"] = {
            "how_can_help": str(raw_interest.get("how_can_help", "")).strip(),
            "weekly_hours": str(raw_interest.get("weekly_hours", "")).strip(),
            "area_interested": [
                str(item).strip()
                for item in raw_interest.get("area_interested", [])
                if str(item).strip()
            ]
            if isinstance(raw_interest.get("area_interested", []), list)
            else [],
        }

    return normalized


def get_member_directory_context() -> dict[str, object]:
    mobile_number = get_authenticated_member_mobile()
    registration = ensure_member_registration(mobile_number)
    family_members: list[dict[str, object]] = []
    for member in list(registration["family_members_details"])[:10]:
        member_copy = dict(member)
        member_copy["calculated_age"] = calculate_age_from_birthdate(str(member.get("birthdate", "")).strip())
        family_members.append(member_copy)
    sangh_activity_interest = normalize_sangh_activity_interest(
        registration["sangh_activity_interest"], family_members
    )
    return {
        "registration": registration,
        "mobile_number": mobile_number,
        "head_of_family": registration["head_of_family_details"],
        "family_members": family_members,
        "family_member_count": len(family_members),
        "sangh_activity_interest": sangh_activity_interest,
        "business_details": registration["business_details"],
    }


def import_member_registrations_from_excel() -> dict[str, int]:
    if not MEMBERS_XLSX_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {MEMBERS_XLSX_PATH}")

    workbook = load_workbook(MEMBERS_XLSX_PATH, read_only=True, data_only=True)
    conn = get_db_connection()
    imported = 0
    skipped = 0
    try:
        worksheet = workbook.active
        conn.execute("BEGIN IMMEDIATE;")
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            part_a = format_excel_part(row[0] if len(row) > 0 else "")
            part_b = format_excel_part(row[1] if len(row) > 1 else "")
            membership_number = f"{part_a}{part_b}".strip()
            name = format_excel_part(row[2] if len(row) > 2 else "")
            father_name = format_excel_part(row[3] if len(row) > 3 else "")
            surname = format_excel_part(row[4] if len(row) > 4 else "")
            full_name = " ".join(part for part in [name, father_name, surname] if part).strip()
            address = format_excel_part(row[5] if len(row) > 5 else "")
            phone = format_excel_part(row[6] if len(row) > 6 else "")
            mobile_number = clean_mobile(row[7] if len(row) > 7 else "")
            email = format_excel_part(row[8] if len(row) > 8 else "")

            if not membership_number or not full_name or not mobile_number:
                skipped += 1
                continue

            head_of_family_details = {
                "full_name": full_name,
                "membership_number": membership_number,
                "sex": "",
                "birthdate": "",
                "mobile_number": mobile_number,
                "phone": phone,
                "blood_group": "",
                "education": "",
                "occupation": "",
                "address": address,
                "city": "",
                "pincode": "",
                "email": email,
                "would_like_to_advertise": "0",
            }

            conn.execute(
                """
                INSERT INTO member_registrations (
                    mobile_number, approval_status, head_of_family_details, updated_at, approved_at
                )
                VALUES (?, 'approved', ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT(mobile_number) DO UPDATE SET
                    head_of_family_details = excluded.head_of_family_details,
                    approval_status = 'approved',
                    updated_at = CURRENT_TIMESTAMP,
                    approved_at = CURRENT_TIMESTAMP
                """,
                (mobile_number, json.dumps(head_of_family_details, ensure_ascii=True)),
            )
            imported += 1

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        workbook.close()
        conn.close()

    return {"imported": imported, "skipped": skipped}


def get_matrimony_profiles() -> list[dict[str, object]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, family_members_details,
                   business_details, created_at, updated_at, approved_at
            FROM member_registrations
            WHERE approval_status = 'approved'
            ORDER BY updated_at DESC, created_at DESC
            """
        ).fetchall()
    finally:
        conn.close()

    profiles: list[dict[str, object]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        business = parse_json_object(row["business_details"])
        family_members = parse_json_array(row["family_members_details"])
        visible_family_members: list[dict[str, str]] = []
        for member in family_members:
            name = str(member.get("name", "")).strip()
            if not name or str(member.get("show_in_matrimony", "0")).strip() != "1":
                continue
            birthdate = str(member.get("birthdate", "")).strip()
            calculated_age = calculate_age_from_birthdate(birthdate)
            member_activity = member.get("sangh_activity_interest", {})
            if not isinstance(member_activity, dict):
                member_activity = {}
            visible_family_members.append(
                {
                    "name": name,
                    "relation": str(member.get("relation", "")).strip(),
                    "sex": str(member.get("sex", "")).strip(),
                    "birthdate": birthdate,
                    "calculated_age": calculated_age,
                    "mobile_number": str(member.get("mobile_number", "")).strip(),
                    "blood_group": str(member.get("blood_group", "")).strip(),
                    "education": str(member.get("education", "")).strip(),
                    "occupation": str(member.get("occupation", "")).strip(),
                    "married": str(member.get("married", "0")).strip(),
                    "mahila_mandal_interest": str(member.get("mahila_mandal_interest", "0")).strip(),
                    "pathshala_interest": str(member.get("pathshala_interest", "0")).strip(),
                    "how_can_help": str(member_activity.get("how_can_help", "")).strip(),
                    "weekly_hours": str(member_activity.get("weekly_hours", "")).strip(),
                    "area_interested": [
                        str(item).strip()
                        for item in member_activity.get("area_interested", [])
                        if str(item).strip()
                    ]
                    if isinstance(member_activity.get("area_interested", []), list)
                    else [],
                }
            )

        if not visible_family_members:
            continue

        profiles.append(
            {
                "id": row["id"],
                "full_name": str(head.get("full_name", "")).strip(),
                "membership_number": str(head.get("membership_number", "")).strip(),
                "address": str(head.get("address", "")).strip(),
                "city": str(head.get("city", "")).strip(),
                "pincode": str(head.get("pincode", "")).strip(),
                "email": str(head.get("email", "")).strip(),
                "mobile_number": row["mobile_number"],
                "business_name": str(business.get("business_name", "")).strip(),
                "business_type": str(business.get("business_type", "")).strip(),
                "designation": str(business.get("designation", "")).strip(),
                "website": str(business.get("website", "")).strip(),
                "notes": str(business.get("notes", "")).strip(),
                "family_members": visible_family_members,
                "registration_id": row["id"],
                "updated_at": format_created_at_ist(row["updated_at"]),
            }
        )
    return profiles


def get_business_search_profiles() -> list[dict[str, str]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT mobile_number, approval_status, head_of_family_details, business_details, updated_at
            FROM member_registrations
            WHERE approval_status = 'approved'
            ORDER BY updated_at DESC
            """
        ).fetchall()
    finally:
        conn.close()

    profiles: list[dict[str, str]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        business = parse_json_object(row["business_details"])
        if str(business.get("show_in_search", "0")).strip() != "1":
            continue

        profiles.append(
            {
                "full_name": str(head.get("full_name", "")).strip(),
                "membership_number": str(head.get("membership_number", "")).strip(),
                "city": str(head.get("city", "")).strip(),
                "business_name": str(business.get("business_name", "")).strip(),
                "business_type": str(business.get("business_type", "")).strip(),
                "designation": str(business.get("designation", "")).strip(),
                "business_contact_mobile": str(business.get("business_contact_mobile", "")).strip() or row["mobile_number"],
                "office_address": str(business.get("office_address", "")).strip(),
                "website": str(business.get("website", "")).strip(),
                "notes": str(business.get("notes", "")).strip(),
                "updated_at": format_created_at_ist(row["updated_at"]),
            }
        )
    return profiles


def get_job_listing_profiles() -> list[dict[str, str]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT mobile_number, head_of_family_details, family_members_details, updated_at
            FROM member_registrations
            WHERE approval_status = 'approved'
            ORDER BY updated_at DESC
            """
        ).fetchall()
    finally:
        conn.close()

    profiles: list[dict[str, str]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        head_name = str(head.get("full_name", "")).strip()
        head_mobile = str(head.get("mobile_number", "")).strip() or row["mobile_number"]
        head_context = {
            "head_of_family": head_name,
            "membership_number": str(head.get("membership_number", "")).strip(),
            "city": str(head.get("city", "")).strip(),
            "address": str(head.get("address", "")).strip(),
            "contact_mobile": head_mobile,
            "updated_at": format_created_at_ist(row["updated_at"]),
        }

        if str(head.get("occupation", "")).strip().lower() == "service":
            profiles.append(
                {
                    **head_context,
                    "name": head_name,
                    "relation": "Head Of Family",
                    "sex": str(head.get("sex", "")).strip(),
                    "birthdate": str(head.get("birthdate", "")).strip(),
                    "mobile_number": head_mobile,
                    "education": str(head.get("education", "")).strip(),
                    "occupation": str(head.get("occupation", "")).strip(),
                    "email": str(head.get("email", "")).strip(),
                }
            )

        for member in parse_json_array(row["family_members_details"]):
            if str(member.get("occupation", "")).strip().lower() != "service":
                continue
            profiles.append(
                {
                    **head_context,
                    "name": str(member.get("name", "")).strip(),
                    "relation": str(member.get("relation", "")).strip(),
                    "sex": str(member.get("sex", "")).strip(),
                    "birthdate": str(member.get("birthdate", "")).strip(),
                    "mobile_number": str(member.get("mobile_number", "")).strip() or head_mobile,
                    "education": str(member.get("education", "")).strip(),
                    "occupation": str(member.get("occupation", "")).strip(),
                    "email": "",
                }
            )
    return profiles


def get_approved_member_contact_numbers() -> set[str]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT mobile_number, head_of_family_details, family_members_details
            FROM member_registrations
            WHERE approval_status = 'approved'
            """
        ).fetchall()
    finally:
        conn.close()

    numbers: set[str] = set()
    for row in rows:
        if cleaned := clean_mobile(row["mobile_number"]):
            numbers.add(cleaned)

        head = parse_json_object(row["head_of_family_details"])
        if cleaned := clean_mobile(head.get("mobile_number", "")):
            numbers.add(cleaned)

        for member in parse_json_array(row["family_members_details"]):
            if cleaned := clean_mobile(member.get("mobile_number", "")):
                numbers.add(cleaned)

    return numbers


def build_member_search_blob(
    head: dict[str, object],
    family_members: list[dict[str, object]],
    sangh_activity_interest: dict[str, object],
    business: dict[str, object],
    mobile_number: str,
) -> str:
    return normalize_text(
        " ".join(
            [
                json.dumps(head, ensure_ascii=True),
                json.dumps(family_members, ensure_ascii=True),
                json.dumps(sangh_activity_interest, ensure_ascii=True),
                json.dumps(business, ensure_ascii=True),
                mobile_number,
            ]
        )
    )


def get_approved_member_search_profiles(query: str = "") -> list[dict[str, object]]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, mobile_number, head_of_family_details, family_members_details,
                   sangh_activity_interest, business_details, updated_at
            FROM member_registrations
            WHERE approval_status = 'approved'
            ORDER BY updated_at DESC, created_at DESC
            """
        ).fetchall()
    finally:
        conn.close()

    normalized_query = normalize_text(query)
    profiles: list[dict[str, object]] = []
    for row in rows:
        head = parse_json_object(row["head_of_family_details"])
        family_members = parse_json_array(row["family_members_details"])
        sangh_activity_interest = parse_json_object(row["sangh_activity_interest"])
        business = parse_json_object(row["business_details"])
        profile = {
            "id": row["id"],
            "full_name": str(head.get("full_name", "")).strip(),
            "membership_number": str(head.get("membership_number", "")).strip(),
            "city": str(head.get("city", "")).strip(),
            "mobile_number": clean_mobile(head.get("mobile_number", "")) or clean_mobile(row["mobile_number"]),
            "family_count": len(family_members),
            "updated_at": format_created_at_ist(row["updated_at"]),
            "search_blob": build_member_search_blob(head, family_members, sangh_activity_interest, business, row["mobile_number"]),
        }
        if normalized_query and normalized_query not in profile["search_blob"]:
            continue
        profiles.append(profile)

    for index, profile in enumerate(profiles, start=1):
        profile["serial_number"] = index
    return profiles


def get_approved_member_profile_detail(registration_id: int) -> dict[str, object] | None:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT id, mobile_number, approval_status, head_of_family_details, family_members_details,
                   sangh_activity_interest, business_details, created_at, updated_at, approved_at
            FROM member_registrations
            WHERE id = ? AND approval_status = 'approved'
            """,
            (registration_id,),
        ).fetchone()
    finally:
        conn.close()

    if not row:
        return None

    head = parse_json_object(row["head_of_family_details"])
    family_members: list[dict[str, object]] = []
    for member in parse_json_array(row["family_members_details"]):
        member_copy = dict(member)
        member_copy["calculated_age"] = calculate_age_from_birthdate(str(member.get("birthdate", "")).strip())
        family_members.append(member_copy)

    return {
        "id": row["id"],
        "mobile_number": clean_mobile(row["mobile_number"]),
        "head_of_family": head,
        "family_members": family_members,
        "sangh_activity_interest": normalize_sangh_activity_interest(
            parse_json_object(row["sangh_activity_interest"]), family_members
        ),
        "business_details": parse_json_object(row["business_details"]),
        "created_at": format_created_at_ist(row["created_at"]),
        "updated_at": format_created_at_ist(row["updated_at"]),
        "approved_at": format_created_at_ist(row["approved_at"]) if row["approved_at"] else "",
    }


def insert_booking(form_data: dict[str, str]) -> int:
    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE;")
        cursor = conn.execute(
            """
            INSERT INTO bookings (booking_date, name, membership_number, murti, address, mobile_number)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                form_data["booking_date"],
                form_data["name"],
                form_data["membership_number"],
                form_data["murti"],
                form_data["address"],
                form_data["mobile_number"],
            ),
        )
        booking_id = int(cursor.lastrowid)
        conn.execute(
            "UPDATE bookings SET booking_number = ? WHERE id = ?",
            (build_booking_number(booking_id), booking_id),
        )
        conn.commit()
        return booking_id
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def update_booking(booking_id: int, form_data: dict[str, str]) -> None:
    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE;")
        conn.execute(
            """
            UPDATE bookings
            SET booking_number = ?, booking_date = ?, name = ?, membership_number = ?, murti = ?, address = ?, mobile_number = ?
            WHERE id = ?
            """,
            (
                form_data["booking_number"],
                form_data["booking_date"],
                form_data["name"],
                form_data["membership_number"],
                form_data["murti"],
                form_data["address"],
                form_data["mobile_number"],
                booking_id,
            ),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def delete_booking(booking_id: int) -> bool:
    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE;")
        cursor = conn.execute("DELETE FROM bookings WHERE id = ?", (booking_id,))
        conn.commit()
        return cursor.rowcount > 0
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def send_confirmation_email(booking: sqlite3.Row) -> None:
    consumer_key = os.getenv("TURBO_SMTP_CONSUMER_KEY")
    consumer_secret = os.getenv("TURBO_SMTP_CONSUMER_SECRET")

    if not consumer_key or not consumer_secret:
        raise RuntimeError(
            "Email is not configured. Set TURBO_SMTP_CONSUMER_KEY and TURBO_SMTP_CONSUMER_SECRET."
        )

    recipients = ["test9@gmail.com", "test@shah.com"]
    mail_from = os.getenv("TURBO_SMTP_FROM", "hello@your-company.com")
    subject = f"Booking Confirmation {booking['booking_number']} - {booking['booking_date']}"
    text_content = "\n".join(
        [
            "Shree Alkpapuri Jain Sangh - Booking Confirmation",
            f"Booking Number: {booking['booking_number']}",
            f"Date: {booking['booking_date']}",
            f"Name: {booking['name']}",
            f"Membership Number: {booking['membership_number']}",
            f"Pratimaji: {booking['murti']}",
            f"Address: {booking['address']}",
            f"Mobile Number: {booking['mobile_number']}",
            f"Submitted At (IST): {booking['created_at']}",
        ]
    )
    html_content = (
        "Shree Alkpapuri Jain Sangh - Booking Confirmation<br>"
        f"<strong>Booking Number:</strong> {booking['booking_number']}<br>"
        f"<strong>Date:</strong> {booking['booking_date']}<br>"
        f"<strong>Name:</strong> {booking['name']}<br>"
        f"<strong>Membership Number:</strong> {booking['membership_number']}<br>"
        f"<strong>Pratimaji:</strong> {booking['murti']}<br>"
        f"<strong>Address:</strong> {booking['address']}<br>"
        f"<strong>Mobile Number:</strong> {booking['mobile_number']}<br>"
        f"<strong>Submitted At (IST):</strong> {booking['created_at']}"
    )

    data = {
        "from": mail_from,
        "to": ",".join(recipients),
        "subject": subject,
        "content": text_content,
        "html_content": html_content,
    }
    data_json = json.dumps(data)
    headers = {
        "Accept": "application/json",
        "Consumerkey": consumer_key,
        "Consumersecret": consumer_secret,
        "Content-Type": "application/json",
    }

    conn = http.client.HTTPSConnection("api.turbo-smtp.com", timeout=30)
    try:
        conn.request("POST", "/api/v2/mail/send", body=data_json, headers=headers)
        response = conn.getresponse()
        response_body = response.read().decode("utf-8", errors="replace")
        if response.status >= 400:
            raise RuntimeError(f"TurboSMTP API error {response.status}: {response_body}")
    finally:
        conn.close()


@app.route("/")
def index():
    return redirect(url_for("member_directory"))

    if request.method == "POST":
        form_data = {
            "booking_date": request.form.get("booking_date", "").strip(),
            "name": request.form.get("name", "").strip(),
            "membership_number": request.form.get("membership_number", "").strip(),
            "murti": request.form.get("murti", "").strip(),
            "address": request.form.get("address", "").strip(),
            "mobile_number": request.form.get("mobile_number", "").strip(),
        }
        member_by_number = None
        by_membership = MEMBER_DIRECTORY.get("by_membership", {})
        if isinstance(by_membership, dict):
            member_by_number = by_membership.get(form_data["membership_number"].upper())
        member_by_name = find_member_by_name(form_data["name"])
        matched_member = member_by_number or member_by_name
        if matched_member:
            if not form_data["membership_number"]:
                form_data["membership_number"] = matched_member["membership_number"]
            if not form_data["address"]:
                form_data["address"] = matched_member["address"]
            if not form_data["mobile_number"]:
                form_data["mobile_number"] = matched_member["mobile_number"]

        if not all(form_data.values()):
            flash("All fields are compulsory.", "error")
            return render_index_page(form_data)

        if not member_by_number:
            flash("Please contact Pedi / Trusti as the membership number not found in the excel member list file.", "error")
            return render_index_page(form_data)

        form_data["membership_number"] = member_by_number["membership_number"]

        if form_data["murti"] not in MURTI_OPTIONS:
            flash("Please select a valid Pratimaji option.", "error")
            return render_index_page(form_data)

        if not re.fullmatch(r"\d{10}", form_data["mobile_number"]):
            flash("Mobile number must be exactly 10 digits.", "error")
            return render_index_page(form_data)

        if not is_otp_verified_for_mobile(form_data["mobile_number"]):
            flash("Please verify OTP for the mobile number before submitting the booking.", "error")
            return render_index_page(form_data)

        try:
            parsed_date = datetime.strptime(form_data["booking_date"], "%Y-%m-%d")
        except ValueError:
            flash("Please provide a valid date.", "error")
            return render_index_page(form_data)

        if parsed_date.date() < today_ist():
            flash("Booking for past dates is not allowed.", "error")
            return render_index_page(form_data)

        if parsed_date.date() > LAST_BOOKING_DATE:
            flash("Booking is allowed only up to 2026-04-30.", "error")
            return render_index_page(form_data)

        if is_tomorrow_booking_cutoff_passed(parsed_date.date()):
            flash("Booking for tomorrow is allowed only before 12:00 PM IST today.", "error")
            return render_index_page(form_data)

        if parsed_date.weekday() == 6:
            flash("Booking is already done", "error")
            return render_index_page(form_data)

        if booking_exists_for_date_and_murti(form_data["booking_date"], form_data["murti"]):
            flash("પ્ણામ, માફ કરજો - તમે આપેલ તારીખમાં તે ભગવાન બીજાએ નોંધાવેલ છે. આપ તે તારીખમાં અન્ય ભગવાન અથવા બીજી તારીખમાં નામ નોંધાવી શકો છો.", "error")
            return render_index_page(form_data)

        try:
            booking_id = insert_booking(form_data)
        except sqlite3.IntegrityError as exc:
            message = str(exc)
            if "UNIQUE constraint failed" in message:
                flash("પ્ણામ, માફ કરજો - તમે આપેલ તારીખમાં તે ભગવાન બીજાએ નોંધાવેલ છે. આપ તે તારીખમાં અન્ય ભગવાન અથવા બીજી તારીખમાં નામ નોંધાવી શકો છો.", "error")
            elif "maximum of two bookings" in message.lower():
                flash("ભાગ્યશાળી, તમે નક્કી કરેલ તારીખમાં બે પ્રતિમાજી નું બુકિંગ થઇ ગયેલ છે. બીજી કોઈ તારીખ નક્કી કરી ફરી પ્રયત્ન કરો.", "error")
            elif "sunday" in message.lower():
                flash("Booking is already done", "error")
            else:
                flash("Could not save booking. Please try again.", "error")
            return render_index_page(form_data)

        clear_otp_session()
        booking = get_booking(booking_id)
        try:
            send_confirmation_email(booking)
            flash("Confirmation email sent successfully.", "success")
        except Exception as exc:
            flash(f"Booking saved, but email could not be sent: {exc}", "error")

        return redirect(url_for("confirmation", booking_id=booking_id))

    clear_otp_session()
    return render_index_page({})


@app.route("/api/send-otp", methods=["POST"])
def send_otp():
    mobile_number = request.form.get("mobile_number", "").strip()
    if not re.fullmatch(r"\d{10}", mobile_number):
        return jsonify({"ok": False, "message": "Please enter a valid 10-digit mobile number."}), 400

    otp = generate_otp()
    expires_at = now_ist() + timedelta(minutes=OTP_EXPIRY_MINUTES)
    try:
        send_otp_sms(mobile_number, otp)
    except Exception as exc:
        return jsonify({"ok": False, "message": f"Could not send OTP: {exc}"}), 500

    session[OTP_SESSION_KEY] = {
        "mobile_number": mobile_number,
        "otp": otp,
        "verified": False,
        "expires_at": expires_at.isoformat(),
    }
    session.modified = True
    return jsonify({"ok": True, "message": "OTP sent successfully."})


@app.route("/api/verify-otp", methods=["POST"])
def verify_otp():
    mobile_number = request.form.get("mobile_number", "").strip()
    otp = request.form.get("otp", "").strip()
    otp_state = get_otp_session()

    if not re.fullmatch(r"\d{10}", mobile_number):
        return jsonify({"ok": False, "message": "Please enter a valid 10-digit mobile number."}), 400
    if not re.fullmatch(r"\d{6}", otp):
        return jsonify({"ok": False, "message": "Please enter the 6-digit OTP."}), 400
    if not otp_state or otp_state.get("mobile_number") != mobile_number:
        return jsonify({"ok": False, "message": "Please send OTP first for this mobile number."}), 400

    try:
        expires_at = datetime.fromisoformat(str(otp_state.get("expires_at", "")))
    except ValueError:
        clear_otp_session()
        return jsonify({"ok": False, "message": "OTP expired. Please send OTP again."}), 400

    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=IST)
    if expires_at < now_ist():
        clear_otp_session()
        return jsonify({"ok": False, "message": "OTP expired. Please send OTP again."}), 400
    if otp_state.get("otp") != otp:
        return jsonify({"ok": False, "message": "Invalid OTP. Please try again."}), 400

    otp_state["verified"] = True
    session[OTP_SESSION_KEY] = otp_state
    session.modified = True
    return jsonify({"ok": True, "message": "OTP verified successfully."})


@app.route("/api/members")
def members_lookup():
    query = request.args.get("q", "").strip()
    return jsonify(search_members(query, limit=25))


@app.route("/api/member-details")
def member_details():
    membership_number = request.args.get("membership_number", "").strip().upper()
    by_membership = MEMBER_DIRECTORY.get("by_membership", {})
    if membership_number and isinstance(by_membership, dict):
        entry = by_membership.get(membership_number)
        if entry:
            return jsonify({"found": True, **entry})

    name = request.args.get("name", "").strip()
    entry_by_name = find_member_by_name(name)
    if entry_by_name:
        return jsonify({"found": True, **entry_by_name})
    return jsonify({"found": False})


@app.route("/member-directory/login")
def member_directory_login():
    if is_member_directory_authenticated():
        return redirect(url_for("member_directory"))
    clear_member_directory_otp_session()
    return render_template("member_directory_login.html", hide_member_directory_hero_extras=True)


@app.route("/api/member-directory/send-otp", methods=["POST"])
def send_member_directory_otp():
    mobile_number = request.form.get("mobile_number", "").strip()
    if not re.fullmatch(r"\d{10}", mobile_number):
        return jsonify({"ok": False, "message": "Please enter a valid 10-digit mobile number."}), 400

    otp = generate_otp()
    expires_at = now_ist() + timedelta(minutes=OTP_EXPIRY_MINUTES)
    try:
        send_otp_sms(mobile_number, otp)
    except Exception as exc:
        return jsonify({"ok": False, "message": f"Could not send OTP: {exc}"}), 500

    session[MEMBER_DIRECTORY_OTP_SESSION_KEY] = {
        "mobile_number": mobile_number,
        "otp": otp,
        "verified": False,
        "expires_at": expires_at.isoformat(),
    }
    session.modified = True
    return jsonify({"ok": True, "message": "OTP sent successfully."})


@app.route("/api/member-directory/verify-otp", methods=["POST"])
def verify_member_directory_otp():
    mobile_number = request.form.get("mobile_number", "").strip()
    otp = request.form.get("otp", "").strip()
    otp_state = get_member_directory_otp_session()

    if not re.fullmatch(r"\d{10}", mobile_number):
        return jsonify({"ok": False, "message": "Please enter a valid 10-digit mobile number."}), 400
    if not re.fullmatch(r"\d{6}", otp):
        return jsonify({"ok": False, "message": "Please enter the 6-digit OTP."}), 400
    if not otp_state or otp_state.get("mobile_number") != mobile_number:
        return jsonify({"ok": False, "message": "Please send OTP first for this mobile number."}), 400

    try:
        expires_at = datetime.fromisoformat(str(otp_state.get("expires_at", "")))
    except ValueError:
        clear_member_directory_otp_session()
        return jsonify({"ok": False, "message": "OTP expired. Please send OTP again."}), 400

    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=IST)
    if expires_at < now_ist():
        clear_member_directory_otp_session()
        return jsonify({"ok": False, "message": "OTP expired. Please send OTP again."}), 400
    if otp_state.get("otp") != otp:
        return jsonify({"ok": False, "message": "Invalid OTP. Please try again."}), 400

    session[MEMBER_DIRECTORY_AUTH_SESSION_KEY] = mobile_number
    clear_member_directory_otp_session()
    ensure_member_registration(mobile_number)
    session.modified = True
    return jsonify({"ok": True, "message": "Login successful."})


@app.route("/qrcode")
@member_directory_required
def qrcode_page():
    mobile_number = session.get(MEMBER_DIRECTORY_AUTH_SESSION_KEY, "")
    registration = ensure_member_registration(str(mobile_number))
    token = build_member_qr_token(int(registration["id"]))
    verification_url = url_for("qrcode_verify", token=token, _external=True)
    return render_template(
        "qrcode.html",
        registration=registration,
        head_of_family=registration["head_of_family_details"],
        verification_url=verification_url,
        fee_status=get_fee_verification_status(registration),
        qr_code_available=qrcode is not None,
        hide_member_directory_hero_extras=True,
    )


@app.route("/qrcode/image.png")
@member_directory_required
def qrcode_image():
    if qrcode is None:
        return "QR code package is not installed on the server.", 503

    mobile_number = session.get(MEMBER_DIRECTORY_AUTH_SESSION_KEY, "")
    registration = ensure_member_registration(str(mobile_number))
    token = build_member_qr_token(int(registration["id"]))
    verification_url = url_for("qrcode_verify", token=token, _external=True)

    image = qrcode.make(verification_url)
    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return send_file(
        output,
        as_attachment=request.args.get("download") == "1",
        download_name=f"member_qrcode_{registration['id']}.png",
        mimetype="image/png",
    )


@app.route("/qrcode/scan")
def qrcode_scan():
    return render_template("qrcode_scan.html", hide_member_directory_hero_extras=True)


@app.route("/qrcode/verify/<token>")
def qrcode_verify(token: str):
    registration_id = get_registration_id_from_qr_token(token)
    registration = get_member_registration_by_id(registration_id) if registration_id else None
    if not registration:
        return render_template(
            "qrcode_verify.html",
            valid=False,
            registration=None,
            head_of_family={},
            fee_status={},
            hide_member_directory_hero_extras=True,
        )

    return render_template(
        "qrcode_verify.html",
        valid=True,
        registration=registration,
        head_of_family=registration["head_of_family_details"],
        fee_status=get_fee_verification_status(registration),
        hide_member_directory_hero_extras=True,
    )


@app.route("/discount-card")
@member_directory_required
def discount_card():
    mobile_number = session.get(MEMBER_DIRECTORY_AUTH_SESSION_KEY, "")
    registration = ensure_member_registration(str(mobile_number))
    token = build_member_qr_token(int(registration["id"]))
    verification_url = url_for("qrcode_verify", token=token, _external=True)
    return render_template(
        "discount_card.html",
        registration=registration,
        head_of_family=registration["head_of_family_details"],
        verification_url=verification_url,
        qr_code_available=qrcode is not None,
        hide_member_directory_hero_extras=True,
    )


@app.route("/discount-card/download")
@member_directory_required
def discount_card_download():
    if qrcode is None:
        flash("QR code package is not installed on the server. Please install updated requirements.txt.", "error")
        return redirect(url_for("discount_card"))

    mobile_number = session.get(MEMBER_DIRECTORY_AUTH_SESSION_KEY, "")
    registration = ensure_member_registration(str(mobile_number))
    head = registration["head_of_family_details"]
    token = build_member_qr_token(int(registration["id"]))
    verification_url = url_for("qrcode_verify", token=token, _external=True)

    qr_image = qrcode.make(verification_url)
    qr_buffer = BytesIO()
    qr_image.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    left = 54
    top = height - 60

    pdf.setTitle("Discount Card")
    pdf.setLineWidth(1.2)
    pdf.roundRect(left, top - 360, width - (left * 2), 360, 16, stroke=1, fill=0)
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(left + 28, top - 38, "Shree Alkapuri Jain Sangh")
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawString(left + 28, top - 66, "Discount Card")

    fields = [
        ("Name", str(head.get("full_name", "")).strip()),
        ("Membership Number", str(head.get("membership_number", "")).strip()),
        ("Mobile Number", str(head.get("mobile_number", "")).strip() or str(registration["mobile_number"])),
        ("Address", str(head.get("address", "")).strip()),
        ("Occupation", str(head.get("occupation", "")).strip()),
        ("Email ID", str(head.get("email", "")).strip()),
    ]

    y = top - 110
    pdf.setFont("Helvetica", 11)
    for label, value in fields:
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(left + 28, y, f"{label}:")
        pdf.setFont("Helvetica", 11)
        text = value or "-"
        if label == "Address" and len(text) > 70:
            first_line = text[:70].rsplit(" ", 1)[0] or text[:70]
            second_line = text[len(first_line):].strip()
            pdf.drawString(left + 150, y, first_line)
            y -= 17
            pdf.drawString(left + 150, y, second_line[:76])
        else:
            pdf.drawString(left + 150, y, text[:82])
        y -= 24

    pdf.drawImage(ImageReader(qr_buffer), width - left - 150, top - 244, width=112, height=112)
    pdf.setFont("Helvetica", 8)
    pdf.drawString(width - left - 166, top - 258, "Scan to verify member and fees status")

    pdf.showPage()
    pdf.save()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"discount_card_{registration['id']}.pdf",
        mimetype="application/pdf",
    )


@app.route("/member-search")
def member_search():
    query = request.args.get("q", "").strip()
    return render_template(
        "member_search.html",
        profiles=get_approved_member_search_profiles(query),
        query=query,
        hide_member_directory_hero_extras=True,
    )


@app.route("/member-search/<int:registration_id>/login")
def member_search_login(registration_id: int):
    profile = get_approved_member_profile_detail(registration_id)
    if not profile:
        flash("Approved member not found.", "error")
        return redirect(url_for("member_search"))
    clear_member_search_otp_session()
    return render_template(
        "member_search_login.html",
        profile=profile,
        registration_id=registration_id,
        hide_member_directory_hero_extras=True,
    )


@app.route("/api/member-search/send-otp", methods=["POST"])
def send_member_search_otp():
    mobile_number = request.form.get("mobile_number", "").strip()
    if not re.fullmatch(r"\d{10}", mobile_number):
        return jsonify({"ok": False, "message": "Please enter a valid 10-digit mobile number."}), 400
    if mobile_number not in get_approved_member_contact_numbers():
        return jsonify({"ok": False, "message": "This mobile number is not allowed for member search access."}), 400

    otp = generate_otp()
    expires_at = now_ist() + timedelta(minutes=OTP_EXPIRY_MINUTES)
    try:
        send_otp_sms(mobile_number, otp)
    except Exception as exc:
        return jsonify({"ok": False, "message": f"Could not send OTP: {exc}"}), 500

    session[MEMBER_SEARCH_OTP_SESSION_KEY] = {
        "mobile_number": mobile_number,
        "otp": otp,
        "expires_at": expires_at.isoformat(),
    }
    session.modified = True
    return jsonify({"ok": True, "message": "OTP sent successfully."})


@app.route("/api/member-search/verify-otp", methods=["POST"])
def verify_member_search_otp():
    mobile_number = request.form.get("mobile_number", "").strip()
    otp = request.form.get("otp", "").strip()
    otp_state = get_member_search_otp_session()

    if not re.fullmatch(r"\d{10}", mobile_number):
        return jsonify({"ok": False, "message": "Please enter a valid 10-digit mobile number."}), 400
    if not re.fullmatch(r"\d{6}", otp):
        return jsonify({"ok": False, "message": "Please enter the 6-digit OTP."}), 400
    if mobile_number not in get_approved_member_contact_numbers():
        return jsonify({"ok": False, "message": "This mobile number is not allowed for member search access."}), 400
    if not otp_state or otp_state.get("mobile_number") != mobile_number:
        return jsonify({"ok": False, "message": "Please send OTP first for this mobile number."}), 400

    try:
        expires_at = datetime.fromisoformat(str(otp_state.get("expires_at", "")))
    except ValueError:
        clear_member_search_otp_session()
        return jsonify({"ok": False, "message": "OTP expired. Please send OTP again."}), 400

    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=IST)
    if expires_at < now_ist():
        clear_member_search_otp_session()
        return jsonify({"ok": False, "message": "OTP expired. Please send OTP again."}), 400
    if otp_state.get("otp") != otp:
        return jsonify({"ok": False, "message": "Invalid OTP. Please try again."}), 400

    session[MEMBER_SEARCH_AUTH_SESSION_KEY] = mobile_number
    clear_member_search_otp_session()
    session.modified = True
    return jsonify({"ok": True, "message": "Login successful."})


@app.route("/member-search/<int:registration_id>")
@member_search_required
def member_search_detail(registration_id: int):
    profile = get_approved_member_profile_detail(registration_id)
    if not profile:
        flash("Approved member not found.", "error")
        return redirect(url_for("member_search"))
    return render_template(
        "member_search_detail.html",
        profile=profile,
        viewer_mobile=get_authenticated_member_search_mobile(),
        hide_member_directory_hero_extras=True,
    )


@app.route("/member-search/logout")
def member_search_logout():
    session.pop(MEMBER_SEARCH_AUTH_SESSION_KEY, None)
    clear_member_search_otp_session()
    return redirect(url_for("member_search"))


@app.route("/member-directory", methods=["GET", "POST"])
@member_directory_required
def member_directory():
    mobile_number = get_authenticated_member_mobile()
    if request.method == "POST":
        section = request.form.get("section", "").strip()
        try:
            if section == "head_of_family":
                payload = parse_head_of_family_form(request.form)
                if not payload["full_name"]:
                    flash("Head of Family name is required.", "error")
                else:
                    update_member_registration_section(mobile_number, "head_of_family_details", payload)
                    flash("Head Of Family Details saved. Status set to pending.", "success")
                    return redirect(url_for("member_directory"))
            elif section == "family_members":
                registration = ensure_member_registration(mobile_number)
                existing_members = list(registration["family_members_details"])[:10]
                if len(existing_members) >= 10:
                    flash("You can add up to 10 family members only.", "error")
                else:
                    payload = parse_family_member_form(request.form)
                    if not payload["name"]:
                        flash("Family member name is required.", "error")
                    else:
                        existing_members.append(payload)
                        update_member_registration_section(mobile_number, "family_members_details", existing_members)
                        flash("Family member details saved. You can add one more member now.", "success")
                        return redirect(url_for("member_directory"))
            elif section == "sangh_activity_interest":
                registration = ensure_member_registration(mobile_number)
                family_members = list(registration["family_members_details"])[:10]
                payload = normalize_sangh_activity_interest(
                    registration["sangh_activity_interest"], family_members
                )
                person_type = request.form.get("person_type", "").strip()
                person_index_raw = request.form.get("person_index", "").strip()
                person_interest = parse_sangh_activity_interest_form(request.form)

                if person_type == "head_of_family":
                    payload["head_of_family"] = person_interest
                elif person_type == "family_member":
                    try:
                        person_index = int(person_index_raw)
                    except ValueError:
                        flash("Invalid family member selected for Sangh Activity.", "error")
                        return redirect(url_for("member_directory"))
                    else:
                        if 0 <= person_index < len(family_members):
                            payload["family_members"][person_index] = person_interest
                        else:
                            flash("Invalid family member selected for Sangh Activity.", "error")
                            return redirect(url_for("member_directory"))
                else:
                    flash("Invalid person selected for Sangh Activity.", "error")
                    return redirect(url_for("member_directory"))

                update_member_registration_section(mobile_number, "sangh_activity_interest", payload)
                flash("Interest in Sangh Activity saved. Status set to pending.", "success")
                return redirect(url_for("member_directory"))
            elif section == "business_details":
                payload = parse_business_details_form(request.form)
                update_member_registration_section(mobile_number, "business_details", payload)
                flash("Business Details saved. Status set to pending.", "success")
                return redirect(url_for("member_directory"))
            else:
                flash("Invalid section submitted.", "error")
        except Exception as exc:
            flash(f"Could not save details: {exc}", "error")

    return render_template(
        "member_directory.html",
        hide_member_directory_hero_extras=True,
        **get_member_directory_context(),
    )


@app.route("/member-directory/logout")
def member_directory_logout():
    session.pop(MEMBER_DIRECTORY_AUTH_SESSION_KEY, None)
    clear_member_directory_otp_session()
    flash("Member Directory logged out.", "success")
    return redirect(url_for("member_directory_login"))


@app.route("/matrimony")
def matrimony():
    return render_template(
        "matrimony.html",
        profiles=get_matrimony_profiles(),
        hide_member_directory_hero_extras=True,
    )


@app.route("/business-search")
def business_search():
    return render_template(
        "business_search.html",
        profiles=get_business_search_profiles(),
        hide_member_directory_hero_extras=True,
    )


@app.route("/job-listing")
def job_listing():
    return render_template(
        "job_listing.html",
        profiles=get_job_listing_profiles(),
        hide_member_directory_hero_extras=True,
    )


@app.route("/new-membership", methods=["GET", "POST"])
def new_membership():
    form_data = {
        "name": "",
        "mobile_number": "",
        "reference_name": "",
        "reference_membership_number": "",
    }
    if request.method == "POST":
        form_data = {
            "name": request.form.get("name", "").strip(),
            "mobile_number": request.form.get("mobile_number", "").strip(),
            "reference_name": request.form.get("reference_name", "").strip(),
            "reference_membership_number": request.form.get("reference_membership_number", "").strip(),
        }
        missing = [label for key, label in {
            "name": "Name",
            "mobile_number": "Mobile",
            "reference_name": "Existing Member Reference Name",
            "reference_membership_number": "Existing Member Membership number",
        }.items() if not form_data[key]]
        if missing:
            flash(f"Please fill: {', '.join(missing)}.", "error")
        elif not re.fullmatch(r"\d{10}", form_data["mobile_number"]):
            flash("Please enter a valid 10-digit mobile number.", "error")
        else:
            insert_new_membership_request(form_data)
            flash("New membership request submitted successfully.", "success")
            return redirect(url_for("new_membership"))

    return render_template(
        "new_membership.html",
        form_data=form_data,
        hide_member_directory_hero_extras=True,
    )


@app.route("/confirmation/<int:booking_id>")
def confirmation(booking_id: int):
    booking = get_booking(booking_id)
    if not booking:
        flash("Booking not found.", "error")
        return redirect(url_for("index"))
    return render_template("confirmation.html", booking=booking)


@app.route("/confirmation/<int:booking_id>/download")
def download_confirmation(booking_id: int):
    booking = get_booking(booking_id)
    if not booking:
        flash("Booking not found.", "error")
        return redirect(url_for("index"))

    content_lines = [
        "Shree Alkpapuri Jain Sangh",
        "Booking Confirmation",
        "",
        f"Booking Number: {booking['booking_number']}",
        f"Date: {booking['booking_date']}",
        f"Name: {booking['name']}",
        f"Membership Number: {booking['membership_number']}",
        f"Pratimaji: {booking['murti']}",
        f"Address: {booking['address']}",
        f"Mobile Number: {booking['mobile_number']}",
        f"Submitted At (IST): {booking['created_at']}",
    ]
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 60
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, y, "Shree Alkpapuri Jain Sangh")
    y -= 24
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, "Booking Confirmation")
    y -= 28
    pdf.setFont("Helvetica", 11)
    for line in content_lines[3:]:
        pdf.drawString(50, y, line)
        y -= 20

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"confirmation_{booking_id}.pdf",
        mimetype="application/pdf",
    )


@app.route("/registration-form.pdf")
def download_registration_form():
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    left = 50
    right = width - 50
    y = height - 55

    pdf.setTitle("Registration Form")
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(left, y, "Shree Alkpapuri Jain Sangh")
    y -= 24
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(left, y, "Registration Form")
    y -= 26

    pdf.setFont("Helvetica", 11)
    pdf.drawString(left, y, "Please fill in the details below.")
    y -= 26

    fields = [
        ("Date", 170),
        ("Name", 360),
        ("Membership Number", 220),
        ("Pratimaji", 240),
        ("Nakro Rakam", 180),
        ("Address", 420),
        ("Mobile Number", 200),
        ("Signature", 180),
    ]

    for label, line_width in fields:
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(left, y, f"{label}:")
        pdf.line(left + 125, y - 2, min(left + 125 + line_width, right), y - 2)
        y -= 34
        if label == "Address":
            pdf.line(left + 125, y - 2, right, y - 2)
            y -= 24
            pdf.line(left + 125, y - 2, right, y - 2)
            y -= 34

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="registration_form.pdf",
        mimetype="application/pdf",
    )


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    next_url = request.args.get("next") or url_for("admin_new_memberships")
    if request.method == "POST":
        password = request.form.get("password", "")
        next_url = request.form.get("next") or url_for("admin_new_memberships")
        if password == ADMIN_PASSWORD:
            session["is_admin_authenticated"] = True
            return redirect(next_url)
        flash("Invalid admin password.", "error")
    return render_template("admin_login.html", next_url=next_url)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin_authenticated", None)
    flash("Admin logged out.", "success")
    return redirect(url_for("index"))


@app.route("/admin")
@admin_required
def admin():
    return redirect(url_for("admin_new_memberships"))


@app.route("/admin/marquee", methods=["POST"])
@admin_required
def admin_update_marquee():
    marquee_text = request.form.get("marquee_text", "").strip()
    update_site_setting("marquee_text", marquee_text)
    flash("Marquee updated successfully.", "success")
    return redirect(request.referrer or url_for("admin"))


@app.route("/admin/members")
@admin_required
def admin_members():
    registrations = get_all_member_registrations()
    return render_template(
        "admin_members.html",
        registrations=registrations,
        hide_member_directory_hero_extras=True,
    )


@app.route("/admin/new-memberships")
@admin_required
def admin_new_memberships():
    return render_template(
        "admin_new_memberships.html",
        requests=get_new_membership_requests(),
        hide_member_directory_hero_extras=True,
    )


@app.route("/admin/advertise-members")
@admin_required
def admin_advertise_members():
    return render_template(
        "admin_advertise_members.html",
        profiles=get_advertise_member_profiles(),
        hide_member_directory_hero_extras=True,
    )


@app.route("/admin/discount-interest-members")
@admin_required
def admin_discount_interest_members():
    return render_template(
        "admin_discount_interest_members.html",
        profiles=get_discount_interest_member_profiles(),
        hide_member_directory_hero_extras=True,
    )


@app.route("/admin/pathshala-interest-members")
@admin_required
def admin_pathshala_interest_members():
    return render_template(
        "admin_pathshala_interest_members.html",
        profiles=get_pathshala_interest_member_profiles(),
        hide_member_directory_hero_extras=True,
    )


@app.route("/admin/sangh-activity-interest")
@admin_required
def admin_sangh_activity_interest():
    selected_area = request.args.get("area", "").strip()
    if selected_area and selected_area not in SANGH_ACTIVITY_AREAS:
        selected_area = ""
    return render_template(
        "admin_sangh_activity_interest.html",
        profiles=get_sangh_activity_interest_profiles(selected_area),
        activity_areas=SANGH_ACTIVITY_AREAS,
        selected_area=selected_area,
        hide_member_directory_hero_extras=True,
    )


@app.route("/admin/members/import", methods=["POST"])
@admin_required
def admin_member_import():
    try:
        result = import_member_registrations_from_excel()
        flash(
            f"Imported {result['imported']} members from ajs.xlsx. Skipped {result['skipped']} rows without membership number, full name, or mobile.",
            "success",
        )
    except Exception as exc:
        flash(f"Could not import member data from ajs.xlsx: {exc}", "error")
    return redirect(url_for("admin_members"))


@app.route("/admin/members/<int:registration_id>/approve", methods=["POST"])
@admin_required
def admin_member_approve(registration_id: int):
    if set_member_registration_approval(registration_id, "approved"):
        flash(f"Member registration {registration_id} approved.", "success")
    else:
        flash("Member registration not found.", "error")
    return redirect(url_for("admin_members"))


@app.route("/admin/members/<int:registration_id>/mark-pending", methods=["POST"])
@admin_required
def admin_member_mark_pending(registration_id: int):
    if set_member_registration_approval(registration_id, "pending"):
        flash(f"Member registration {registration_id} marked as pending.", "success")
    else:
        flash("Member registration not found.", "error")
    return redirect(url_for("admin_members"))


@app.route("/admin/members/<int:registration_id>/fees", methods=["POST"])
@admin_required
def admin_member_add_fee(registration_id: int):
    year = request.form.get("year", "").strip()
    amount = request.form.get("amount", "").strip()
    try:
        updated = add_member_registration_fee(registration_id, year, amount)
    except ValueError as exc:
        flash(str(exc), "error")
    else:
        if updated:
            flash(f"Fee added for member registration {registration_id}.", "success")
        else:
            flash("Member registration not found.", "error")
    return redirect(url_for("admin_members"))


@app.route("/admin/edit/<int:booking_id>", methods=["GET", "POST"])
@admin_required
def admin_edit(booking_id: int):
    booking = get_booking_for_edit(booking_id)
    if not booking:
        flash("Booking not found.", "error")
        return redirect(url_for("admin"))

    form_data = {
        "booking_number": booking["booking_number"],
        "booking_date": booking["booking_date"],
        "name": booking["name"],
        "membership_number": booking["membership_number"],
        "murti": booking["murti"],
        "address": booking["address"],
        "mobile_number": booking["mobile_number"],
    }

    if request.method == "POST":
        form_data = {
            "booking_number": request.form.get("booking_number", "").strip(),
            "booking_date": request.form.get("booking_date", "").strip(),
            "name": request.form.get("name", "").strip(),
            "membership_number": request.form.get("membership_number", "").strip(),
            "murti": request.form.get("murti", "").strip(),
            "address": request.form.get("address", "").strip(),
            "mobile_number": request.form.get("mobile_number", "").strip(),
        }
        edit_password = request.form.get("edit_password", "")
        if edit_password != ADMIN_EDIT_PASSWORD:
            flash("Invalid edit password.", "error")
            return render_template(
                "admin_edit.html",
                booking=booking,
                form_data=form_data,
                murtis=MURTI_OPTIONS,
                last_booking_date=LAST_BOOKING_DATE.isoformat(),
            )

        member_by_number = None
        by_membership = MEMBER_DIRECTORY.get("by_membership", {})
        if isinstance(by_membership, dict):
            member_by_number = by_membership.get(form_data["membership_number"].upper())

        if not all(form_data.values()):
            flash("All fields are compulsory.", "error")
        elif booking_number_exists_excluding_id(form_data["booking_number"], booking_id):
            flash("Booking Number already exists. Please use a unique Booking Number.", "error")
        elif not member_by_number:
            flash("Please contact Pedi / Trusti as the membership number not found in the excel member list file.", "error")
        elif form_data["murti"] not in MURTI_OPTIONS:
            flash("Please select a valid Pratimaji option.", "error")
        elif not re.fullmatch(r"\d{10}", form_data["mobile_number"]):
            flash("Mobile number must be exactly 10 digits.", "error")
        else:
            try:
                parsed_date = datetime.strptime(form_data["booking_date"], "%Y-%m-%d")
            except ValueError:
                parsed_date = None
                flash("Please provide a valid date.", "error")

            if parsed_date:
                if parsed_date.date() > LAST_BOOKING_DATE:
                    flash("Booking is allowed only up to 2026-04-30.", "error")
                elif parsed_date.weekday() == 6 and booking["is_dummy"] != "1":
                    flash("Booking is already done", "error")
                elif booking_exists_for_date_and_murti_excluding_id(form_data["booking_date"], form_data["murti"], booking_id):
                    flash("પ્ણામ, માફ કરજો - તમે આપેલ તારીખમાં તે ભગવાન બીજાએ નોંધાવેલ છે. આપ તે તારીખમાં અન્ય ભગવાન અથવા બીજી તારીખમાં નામ નોંધાવી શકો છો.", "error")
                elif count_bookings_for_date_excluding_id(form_data["booking_date"], booking_id) >= 2:
                    flash("ભાગ્યશાળી, તમે નક્કી કરેલ તારીખમાં બે પ્રતિમાજી નું બુકિંગ થઇ ગયેલ છે. બીજી કોઈ તારીખ નક્કી કરી ફરી પ્રયત્ન કરો.", "error")
                else:
                    form_data["membership_number"] = member_by_number["membership_number"]
                    try:
                        update_booking(booking_id, form_data)
                        flash(f"Booking {booking_id} updated successfully.", "success")
                        return redirect(url_for("admin"))
                    except sqlite3.IntegrityError as exc:
                        message = str(exc)
                        if "UNIQUE constraint failed" in message:
                            flash("પ્ણામ, માફ કરજો - તમે આપેલ તારીખમાં તે ભગવાન બીજાએ નોંધાવેલ છે. આપ તે તારીખમાં અન્ય ભગવાન અથવા બીજી તારીખમાં નામ નોંધાવી શકો છો.", "error")
                        else:
                            flash("Could not update booking. Please try again.", "error")

        booking = {**booking, **form_data}

    return render_template(
        "admin_edit.html",
        booking=booking,
        form_data=form_data,
        murtis=MURTI_OPTIONS,
        last_booking_date=LAST_BOOKING_DATE.isoformat(),
    )


@app.route("/admin/export.xlsx")
@admin_required
def admin_export():
    filters = normalize_filters(request.args)
    sort_by, sort_dir = normalize_admin_sort(request.args)
    bookings = get_admin_bookings(filters, sort_by, sort_dir)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Bookings"
    headers = [
        "ID",
        "Booking Number",
        "Date",
        "Name",
        "Membership Number",
        "Pratimaji",
        "Address",
        "Mobile Number",
        "Created At (IST)",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for booking in bookings:
        worksheet.append(
            [
                booking["id"],
                booking["booking_number"],
                f"{booking['booking_date_display']} ({booking['weekday']})".strip(),
                booking["name"],
                booking["membership_number"],
                booking["murti"],
                booking["address"],
                booking["mobile_number"],
                booking["created_at"],
            ]
        )

    for col in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        worksheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="bookings.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/delete/<int:booking_id>", methods=["POST"])
@admin_required
def admin_delete(booking_id: int):
    password = request.form.get("delete_password", "")
    if password != ADMIN_DELETE_PASSWORD:
        flash("Invalid delete password.", "error")
        return redirect(url_for("admin"))

    if delete_booking(booking_id):
        flash(f"Booking {booking_id} deleted successfully.", "success")
    else:
        flash("Booking not found.", "error")
    return redirect(url_for("admin"))


init_db()
load_member_directory()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
